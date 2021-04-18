import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

from .column_sets import ColumnFormats


class ReportRenderer:
    def render(self, json_table):
        return json_table


class JsonRenderer(ReportRenderer):
    def render(self, saved_report):
        columns = self.get_columns(saved_report)
        return {"columns": columns, "dataSource": saved_report.generate()}

    def get_columns(self, saved_report):
        columnsets = saved_report.report.get_usable_columnsets(
            saved_report.columns_data
        )
        return [
            {
                "dataIndex": columnset.key,
                "key": columnset.key,
                "title": columnset.label,
                "children": [
                    self.get_column(columnset_item, saved_report)
                    for columnset_item in columnset.items
                ],
            }
            for columnset in columnsets
        ]

    def get_column(self, columnset_item, saved_report):
        column = saved_report.report.get_column_by_key(columnset_item.column_key)
        return {
            "dataIndex": column.key,
            "key": column.key,
            "title": column.label,
        }


class ExcelRenderer(ReportRenderer):
    def render(self, saved_report):
        self.saved_report = saved_report
        data = saved_report.generate()

        self.report = saved_report.report

        wb = Workbook()
        self.ws = wb.active
        columnsets = self.report.get_usable_columnsets(saved_report.columns_data)
        curr_col = 1
        curr_line = 1
        # if saved_report.name:
        #     self.render_title(saved_report.name, columnsets)
        #     curr_line = 2
        for columnset in columnsets:
            curr_col = self.render_columnset(columnset, curr_col, curr_line)
        curr_line = curr_line + 2
        for line in data:
            curr_line = self.render_line(line, curr_line)
        for i in range(40):
            self.ws.column_dimensions[get_column_letter(i + 1)].width = 15
        for i in range(1, 40):
            self.ws.row_dimensions[i].height = 35
        stream = io.BytesIO()
        wb.save(stream)
        return io.BytesIO(stream.getvalue())

    def render_title(self, title, columnsets):
        dimension = sum(len(columnset.items) for columnset in columnsets)
        self.ws["A1"] = title
        self.format_title(self.ws["A1"])
        self.ws.merge_cells(f"A1:{get_column_letter(dimension)}1")

    def render_columnset(self, columnset, curr_col, curr_line):
        label_start = f"{get_column_letter(curr_col)}{curr_line}"
        col_end = curr_col + len(columnset.items) - 1
        label_end = f"{get_column_letter(col_end)}1"

        self.ws[label_start] = columnset.label
        self.format_title(self.ws[label_start])
        self.ws.merge_cells(f"{label_start}:{label_end}")
        for i, item in enumerate(columnset.items):
            cell = f"{get_column_letter(curr_col + i)}{curr_line + 1}"
            self.ws[cell] = self.report.get_column_by_key(item.column_key).label
            self.format_cell(self.ws[cell], item)
        return curr_col + len(columnset.items)

    def render_line(self, line, n):
        i = 1
        for columnset in self.report.get_usable_columnsets(
            self.saved_report.columns_data
        ):
            for item in columnset.items:
                cell = f"{get_column_letter(i)}{n}"
                self.ws[cell] = line[item.column_key]
                self.format_cell(self.ws[cell], item)
                i += 1
        return n + 1

    def format_title(self, cell):
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def format_cell(self, cell, columnset_item):
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        if ColumnFormats.emphasis in columnset_item.formats:
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
